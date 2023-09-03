'''
读取result.txt和pos.txt文件中的字符串和坐标 并将其按照序号合并到一个output.txt文件中
TiltCorrectionOutput.txt文件中的每一行按照左上角坐标从上到下、从左到右排序
'''
import openpyxl
# 读取第一个文件中的中文字符串
with open(r'C:\Users\Alex\PycharmProjects\Table_rec_system\Table_detection\results\Identification_res.txt', 'r') as f:
    lines = f.readlines()
    strings = {}
    for line in lines:
        parts = line.strip().split()
        if len(parts) == 2:
            index, string = parts
            strings[index] = string

# 读取第二个文件中的坐标
with open(r'C:\Users\Alex\PycharmProjects\Table_rec_system\Table_detection\results\pos_res.txt', 'r') as f:
    lines = f.readlines()
    coords = {}
    for line in lines:
        parts = line.strip().split()
        if len(parts) == 2:
            index, coord = parts
            coords[index] = coord

results = []
# 匹配中文字符串和坐标并输出到新文件
with open('output.txt', 'w', encoding='utf-8') as f:
    for index in strings:
        if index in coords:
            string = strings[index]
            coord = coords[index]
            results.append((string, coord))

# 转换数据格式
new_data = []
for item in results:
    index = item[0]
    coords = item[1].split(',')
    x1, y1, x2, y2 = map(int, coords)
    new_data.append((index, (x1, y1), (x2, y2)))
new_data.sort(key=lambda x: (x[1][1], x[1][0]))
#print(new_data)

# 输出排序后的结果到新文件
with open('output1.txt', 'w', encoding='utf-8') as f:
    for result in new_data:
        string, coordl, coordr = result
        f.write(f"{string} {coordl} {coordr}\n")

# 定义纵坐标的容差范围
row_tolerance = 10
#print(new_data)
# 计算表格的行数
row_indices = [1]
for i in range(1, len(new_data)):
    prev_coord = new_data[i-1][1]
    curr_coord = new_data[i][1]
    if i+1 < len(new_data):
        next_coord = new_data[i+1][1]
        if abs(curr_coord[1] - prev_coord[1]) > row_tolerance:
            row_indices.append(i + 1)
rows = len(row_indices)
# print(row_indices)
# print(rows)

# 计算每个单元格的高度和宽度
cell_sizes = {}
for item in new_data:
    string, coord1, coord2 = item
    width = coord2[0] - coord1[0]
    height = coord2[1] - coord1[1]
    width = (width % 100) % 45
    height = 20
    #print(width, height)
    cell_sizes[string] = (width, height)

# 创建一个新的 Excel 工作簿
wb = openpyxl.Workbook()
ws = wb.active

#将结果填入到单元格当中
row_index = 1
col_index = 1
index = 0
while row_index <= rows:
    while col_index < row_indices[row_index]:
        if index < len(new_data):
            #print(row_index, col_index - row_indices[row_index - 1] + 1)
            # if row_index == 1:
            #     ws.cell(row=row_index, column=col_index).value = new_data[index][0]
            # else:
            ws.cell(row=row_index, column=col_index - row_indices[row_index - 1] + 1).value = new_data[index][0]
            # 修改单元格的大小
            if string in cell_sizes:
                width, height = cell_sizes[string]
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_index - row_indices[row_index - 1] + 1)].width = width
                ws.row_dimensions[row_index].height = height

            index += 1
            col_index += 1
    row_index += 1

#对最后一行进行特殊处理
    if row_index == len(row_indices):
        new_col_index = 1
        while index < len(new_data):
            ws.cell(row=row_index, column=new_col_index).value = new_data[index][0]
            if string in cell_sizes:
                width, height = cell_sizes[string]
                ws.column_dimensions[openpyxl.utils.get_column_letter(new_col_index)].width = width
                ws.row_dimensions[row_index].height = height

            index += 1
            new_col_index += 1
        break

wb.save('FinalExcel.xlsx')
print("The doc and excel have been generated")





