from openpyxl import load_workbook
from docx import Document

# 加载 Excel 文件
wb = load_workbook('FinalExcel.xlsx')
ws = wb.active

# 定义要复制的 Excel 区域
excel_start_row = 3
excel_end_row = 24
excel_start_col = 1
excel_end_col = 10

# 定义要粘贴到的 Word 表格位置
word_start_row = 2
word_start_col = 0

# 加载 Word 文档并获取现有表格
doc = Document('008.docx')
table = doc.tables[0]

# 将 Excel 数据写入 Word 表格
for row_index, row in enumerate(ws.iter_rows(min_row=excel_start_row, max_row=excel_end_row, min_col=excel_start_col, max_col=excel_end_col)):
    for col_index, cell in enumerate(row):
        table.cell(word_start_row + row_index, word_start_col + col_index).text = str(cell.value)

# 保存 Word 文档
doc.save('FinalDoc.docx')
